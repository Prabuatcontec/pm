# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""
from flask import jsonify, session, make_response, Response
from apps.home import blueprint

from flask_login import login_required
from apps.report.models import motions

from datetime import datetime, timedelta
import time, calendar
import openpyxl, os
from openpyxl.styles import Color, PatternFill, Font, Border
import pickle

stationDis = {"Line1Station": "CLD-SHIP13", "Line2Station": "CLD-SHIP15", "Line3Station": "CLD-SHIP16",
              "Line4Station": "CLD-SHIP21"}





@blueprint.route('/report/data/<startdate>/<enddate>/<day>')
@login_required
def reportdatesearch(startdate,enddate,day):
    stationTime = motions.motion_loader_byarea_date(session['search_station'], startdate, enddate,day)
    time_report_count = {}
    time_report = {}
    time_report_hrs = {}
    time_report_time = {}
    pretime = {}

    station_all_day = {}
    for station in stationTime:
        startTime = station[0]
        endTime = station[4]
        stationstarttime = session['timedep'] + station[0]
        stationendtime = session['timedep'] + station[4]
        current_time = datetime.fromtimestamp(stationstarttime).strftime('%H:%M:%S')

        dt_obj = datetime.fromtimestamp(stationstarttime).strftime('%d-%m-%Y')
        # print(  current_time)
        if is_between(current_time, ("07:00:00", "20:00:00")):
            if dt_obj not in time_report_count:
                time_report_count[dt_obj] = {"1-2": 0, "2-3": 0, "3-5": 0, "5-10": 0, "10-15": 0, "15-60": 0}
                time_report[dt_obj] = {"1-2": 0, "2-3": 0, "3-5": 0, "5-10": 0, "10-15": 0, "15-60": 0}
                time_report_hrs[dt_obj] = {"1-2": 0, "2-3": 0, "3-5": 0, "5-10": 0, "10-15": 0, "15-60": 0}
                time_report_time[dt_obj] = {"1-2": [], "2-3": [], "3-5": [], "5-10": [], "10-15": [], "15-60": []}
                pretime[dt_obj] = 0
            if pretime[dt_obj] > int(stationstarttime) or pretime[dt_obj] == 0:
                pretime[dt_obj] = int(stationstarttime)
             
            
            

            

            

            if 900 < station[5] <= 1000:
                p_time = time_report[dt_obj]["15-60"]
                time_report_count[dt_obj]["15-60"] = int(time_report_count[dt_obj]["15-60"]) + 1
                time_report[dt_obj]["15-60"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["15-60"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["15-60"].append({"from": datetime.fromtimestamp(int(stationstarttime)).strftime("%Y-%m-%d %H:%M:%S"),
                                                          "to": datetime.fromtimestamp(int(stationendtime)).strftime("%Y-%m-%d %H:%M:%S"),
                                                          "diff": int(int(station[5])/60),
                                                          "startTime":startTime,
                                                          "endTime":endTime})
            if 600 < station[5] <= 900:
                p_time = time_report[dt_obj]["10-15"]
                time_report_count[dt_obj]["10-15"] = int(time_report_count[dt_obj]["10-15"]) + 1
                time_report[dt_obj]["10-15"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["10-15"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["10-15"].append({"from": datetime.fromtimestamp(int(stationstarttime)).strftime("%Y-%m-%d %H:%M:%S"),
                                                          "to": datetime.fromtimestamp(int(stationendtime)).strftime("%Y-%m-%d %H:%M:%S"),
                                                          "diff": int(int(station[5])/60),
                                                          "startTime":startTime,
                                                          "endTime":endTime})

            if 300 < station[5] <= 600:
                p_time = time_report[dt_obj]["5-10"]
                time_report_count[dt_obj]["5-10"] = int(time_report_count[dt_obj]["5-10"]) + 1
                time_report[dt_obj]["5-10"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["5-10"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["5-10"].append({"from": datetime.fromtimestamp(int(stationstarttime)).strftime("%Y-%m-%d %H:%M:%S"),
                                                         "to": datetime.fromtimestamp(int(stationendtime)).strftime("%Y-%m-%d %H:%M:%S"),
                                                         "diff": int(int(station[5])/60),
                                                          "startTime":startTime,
                                                          "endTime":endTime})
            if 180 < station[5] <= 300:
                p_time = time_report[dt_obj]["3-5"]
                time_report_count[dt_obj]["3-5"] = int(time_report_count[dt_obj]["3-5"]) + 1
                time_report[dt_obj]["3-5"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["3-5"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["3-5"].append({"from": datetime.fromtimestamp(int(stationstarttime)).strftime("%Y-%m-%d %H:%M:%S"),
                                                        "to": datetime.fromtimestamp(int(stationendtime)).strftime("%Y-%m-%d %H:%M:%S"),
                                                        "diff": int(int(station[5])/60),
                                                          "startTime":startTime,
                                                          "endTime":endTime})
            if 120 < station[5] <= 180:
                p_time = time_report[dt_obj]["2-3"]
                time_report_count[dt_obj]["2-3"] = int(time_report_count[dt_obj]["2-3"]) + 1
                time_report[dt_obj]["2-3"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["2-3"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["2-3"].append({"from": datetime.fromtimestamp(int(stationstarttime)).strftime("%Y-%m-%d %H:%M:%S"),
                                                        "to": datetime.fromtimestamp(int(stationendtime)).strftime("%Y-%m-%d %H:%M:%S"),
                                                        "diff": int(int(station[5])/60),
                                                          "startTime":startTime,
                                                          "endTime":endTime})

    # print(time_report_time)
    station_all_day = {"time_report_count": time_report_count, "time_report": time_report,
                       "time_report_hrs": time_report_hrs, "pretime": pretime,"time_report_time":time_report_time}

    result = {'result': station_all_day}

    
    return jsonify(result), 200



@blueprint.route('/report/box')
@login_required
def report_box():
    stationBox = {"Line1Station": "box1", "Line2Station": "box1", "Line3Station": "box2", "Line4Station": "box3",
                  "Line5Station": "box4"}
    stationTime = motions.motion_loader_byarea(stationBox[session['search_station']], '1', '1', 30, 5)

    last_week_count = 0
    stationTimeData = motions.get_cnt_lastweek(stationDis[session['search_station']])
    for station in stationTimeData:
        last_week_count = station[0]

    time_report_count = {}
    time_report = {}
    time_report_hrs = {}

    for station in stationTime:
        stationstarttime = session['timedep'] + station[0]
        current_time = datetime.fromtimestamp(stationstarttime).strftime('%H:%M:%S')

        dt_obj = datetime.fromtimestamp(stationstarttime).strftime('%d-%m-%Y')
        if is_between(current_time, ("07:00:00", "20:00:00")):
            if dt_obj not in time_report_count:
                time_report_count[dt_obj] = {"1-2": 0}
                time_report[dt_obj] = {"1-2": 0}
                time_report_hrs[dt_obj] = {"1-2": 0}

            if station[5] > 0:
                p_time = time_report[dt_obj]["1-2"]
                time_report_count[dt_obj]["1-2"] = int(time_report_count[dt_obj]["1-2"]) + 1
                time_report[dt_obj]["1-2"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["1-2"] = convert_time((int(p_time) + station[5]))

    station_all_day = {"time_report_count": time_report_count, "time_report": time_report,
                       "time_report_hrs": time_report_hrs, "station_name": session['search_station'],
                       "last_week_count": last_week_count}

    result = {'result': station_all_day}

    return jsonify(result), 200


@blueprint.route('/report/shipping')
@login_required
def report_shipping():
    stationTime = motions.actionin_shipping_data(stationDis[session['search_station']])
    time_report_count = {}
    time_report = {}
    time_report_hrs = {}
    pretime = {}
    hrShippingCount = {}
    customers = {}

    shipping_cus_count = {}

    warehouses = {}

    now = datetime.now()
    for x in range(8):
        d = now - timedelta(days=x)
        dt_objr = d.strftime("%d-%m-%Y")

        if dt_objr not in shipping_cus_count:
            shipping_cus_count[dt_objr] = {}

    for station in stationTime:
        stationstarttime = session['timedep'] + station[3]

        dt_obj = datetime.fromtimestamp(stationstarttime).strftime('%d-%m-%Y')

        str_customer = str(station[6])[0:3]
        wh = station[7].strip()
        if str_customer not in customers:
            customers[str_customer] = str_customer
        if wh not in warehouses:
            warehouses = wh

        if str_customer not in shipping_cus_count[dt_obj]:
            shipping_cus_count[dt_obj][str_customer] = 0
        shipping_cus_count[dt_obj][str_customer] = shipping_cus_count[dt_obj][str_customer] + 1

        if dt_obj not in time_report_count:
            time_report_count[dt_obj] = {"0-1": 0, "1-2": 0, "2-3": 0, "3-5": 0, "5-10": 0, "10-15": 0, "15-60": 0}
            time_report[dt_obj] = {"0-1": 0, "1-2": 0, "2-3": 0, "3-5": 0, "5-10": 0, "10-15": 0, "15-60": 0}
            time_report_hrs[dt_obj] = {"0-1": 0, "1-2": 0, "2-3": 0, "3-5": 0, "5-10": 0, "10-15": 0, "15-60": 0}
            pretime[dt_obj] = 0
            hrShippingCount[dt_obj] = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0,
                                       8: 0, 9: 0, 10: 0, 11: 0, 12: 0, 13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0,
                                       20: 0,
                                       21: 0, 22: 0, 23: 0}
        current_time = datetime.fromtimestamp(stationstarttime).strftime('%H')

        hrShippingCount[dt_obj][int(current_time)] = hrShippingCount[dt_obj][int(current_time)] + 1
        if pretime[dt_obj] > int(stationstarttime) or pretime[dt_obj] == 0:
            pretime[dt_obj] = int(stationstarttime)
        if int(station[2]) < 60:
            p_time = time_report[dt_obj]["0-1"]
            time_report_count[dt_obj]["0-1"] = int(time_report_count[dt_obj]["0-1"]) + 1
            time_report[dt_obj]["0-1"] = (int(p_time) + int(station[2]))
            time_report_hrs[dt_obj]["0-1"] = convert_time((int(p_time) + int(station[2])))

        if 59 < int(station[2]) <= 120:
            p_time = time_report[dt_obj]["1-2"]
            time_report_count[dt_obj]["1-2"] = int(time_report_count[dt_obj]["1-2"]) + 1
            time_report[dt_obj]["1-2"] = (int(p_time) + int(station[2]))
            time_report_hrs[dt_obj]["1-2"] = convert_time((int(p_time) + int(station[2])))
        if 120 < int(station[2]) <= 180:
            p_time = time_report[dt_obj]["2-3"]
            time_report_count[dt_obj]["2-3"] = int(time_report_count[dt_obj]["2-3"]) + 1
            time_report[dt_obj]["2-3"] = (int(p_time) + int(station[2]))
            time_report_hrs[dt_obj]["2-3"] = convert_time((int(p_time) + int(station[2])))

        if 180 < int(station[2]) <= 300:
            p_time = time_report[dt_obj]["3-5"]
            time_report_count[dt_obj]["3-5"] = int(time_report_count[dt_obj]["3-5"]) + 1
            time_report[dt_obj]["3-5"] = (int(p_time) + int(station[2]))
            time_report_hrs[dt_obj]["3-5"] = convert_time((int(p_time) + int(station[2])))

        if 300 < int(station[2]) <= 600:
            p_time = time_report[dt_obj]["5-10"]
            time_report_count[dt_obj]["5-10"] = int(time_report_count[dt_obj]["5-10"]) + 1
            time_report[dt_obj]["5-10"] = (int(p_time) + int(station[2]))
            time_report_hrs[dt_obj]["5-10"] = convert_time((int(p_time) + int(station[2])))

        if 600 < int(station[2]) <= 900:
            p_time = time_report[dt_obj]["10-15"]
            time_report_count[dt_obj]["10-15"] = int(time_report_count[dt_obj]["10-15"]) + 1
            time_report[dt_obj]["10-15"] = (int(p_time) + int(station[2]))
            time_report_hrs[dt_obj]["10-15"] = convert_time((int(p_time) + int(station[2])))

        if 900 < int(station[2]) <= 3600:
            p_time = time_report[dt_obj]["15-60"]
            time_report_count[dt_obj]["15-60"] = int(time_report_count[dt_obj]["15-60"]) + 1
            time_report[dt_obj]["15-60"] = (int(p_time) + int(station[2]))
            time_report_hrs[dt_obj]["15-60"] = convert_time((int(p_time) + int(station[2])))

    station_all_day = {"time_report_count": time_report_count,
                       "time_report": time_report,
                       "time_report_hrs": time_report_hrs,
                       "pretime": pretime,
                       "hrShippingCount": hrShippingCount,
                       "shipping_cus_count": shipping_cus_count,
                       "warehouses": warehouses,
                       "customers": customers}

    pickle.dump(time_report_hrs, open(get_correct_path(session['search_station'] + "_hr_Shipping.p"), "wb"))
    pickle.dump(hrShippingCount, open(get_correct_path(session['search_station'] + "_hrShippingCount.p"), "wb"))
    result = {'result': station_all_day}
    return jsonify(result), 200


@blueprint.route('/report-csv')
@login_required
def report_csv():
    now = datetime.now()

    rows = ["1-2", "2-3", "3-5", "5-10", "10-15", "15-60"]
    time_report_count_csv = get_station_data(session['search_station'] + "_data_count.p")

    data = [["Operator Activity"]]
    data.append(["Date", "1-2 Min", "2-3 Min", "3-5 Min", "5-10 Min", "10-15 Min", "15-60 Min"])
    # print(time_report_count_csv)
    for x in range(7):
        d = now - timedelta(days=x)

        if d.strftime("%d-%m-%Y") not in time_report_count_csv:
            time_report_count_csv[d.strftime("%d-%m-%Y")] = {"1-2": 0, "2-3": 0,
                                                             "3-5": 0, "5-10": 0, "10-15": 0,
                                                             "15-60": 0}
        for i, val in enumerate(rows):
            if val not in time_report_count_csv[d.strftime("%d-%m-%Y")]:
                time_report_count_csv[d.strftime("%d-%m-%Y")][val] = 0
        dataAdd = [d.strftime("%d-%m-%Y"), str(time_report_count_csv[d.strftime("%d-%m-%Y")]["1-2"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["2-3"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["3-5"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["5-10"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["10-15"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["15-60"])]

        data.append(dataAdd)

    time_report_time = get_station_data(session['search_station'] + "_data.p")
    data.append([])
    data.append([])
    data.append(["Operator Activity (From - To)"])
    lo_arr = []
    lo_arr_in = []
    vo = 13
    for x in range(7):
        d = now - timedelta(days=x)
        if d.strftime("%d-%m-%Y") not in time_report_time:
            time_report_time[d.strftime("%d-%m-%Y")] = {"1-2": [], "2-3": [],
                                                        "3-5": [], "5-10": [], "10-15": [],
                                                        "15-60": []}

        for i, val in enumerate(rows):
            if val not in time_report_time[d.strftime("%d-%m-%Y")]:
                time_report_time[d.strftime("%d-%m-%Y")][val] = []
        lo_arr.append(vo)

        data.append([d.strftime("%d-%m-%Y")])
        vo = vo + 1
        for o, datapop in enumerate(rows):
            data.append([datapop + ' Min'])
            lo_arr_in.append(vo)
            vo = vo + 1
            for i, val in enumerate(time_report_time[d.strftime("%d-%m-%Y")][datapop]):
                data.append([str(val["from"]) + " - " + str(val["to"]), str(val["diff"])])
                vo = vo + 1
            data.append([])
            vo = vo + 1

    fname2 = r'Operator_Activity' + '_' + session['_user_id'] + '.xlsx'
    workbook = openpyxl.Workbook(fname2)
    if os.path.isfile(fname2):
        print('old file')
    else:
        print('new file')
        workbook2 = openpyxl.Workbook(fname2)
        workbook2.save(fname2)
    wb = openpyxl.load_workbook(fname2)
    ws = wb.active

    sheet = wb.active

    for row in data:
        sheet.append(row)
    redFill = PatternFill(start_color='FFFF00',
                          end_color='FFFF99',
                          fill_type='solid')

    redFill_in = PatternFill(start_color='A9A9A9',
                             end_color='E8E8E8',
                             fill_type='solid')

    for cell in ws["2:2"]:
        cell.fill = redFill
        cell.font = Font(size="14", color='484848')
    for p, pon in enumerate(lo_arr):
        for cell in ws[str(pon) + ":" + str(pon)]:
            cell.fill = redFill
            cell.font = Font(size="14", color='484848')
    for p, pon in enumerate(lo_arr_in):
        for cell in ws[str(pon) + ":" + str(pon)]:
            cell.fill = redFill_in
            cell.font = Font(size="14", color='FFFFFF')

    fontStyle = Font(size="20", color='484848')
    ws['A1'].font = fontStyle

    ws['A12'].font = fontStyle

    wb.save(fname2)
    excelDownload = open(fname2, 'rb').read()
    os.remove(fname2)
    return Response(
        excelDownload,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition":
                     "attachment; filename=" + fname2})


@blueprint.route('/report-shipping')
@login_required
def report_shipping_hr():
    now = datetime.now()

    rows = ["0-1", "1-2", "2-3", "3-5", "5-10", "10-15", "15-60"]
    time_report_count_csv = get_station_data(session['search_station'] + "_hr_Shipping.p")

    data = [["Shipping Activity"]]
    data.append(["Date", "0-1 Min", "1-2 Min", "2-3 Min", "3-5 Min", "5-10 Min", "10-15 Min", "15-60 Min"])
    # print(time_report_count_csv)
    for x in range(7):
        d = now - timedelta(days=x)

        if d.strftime("%d-%m-%Y") not in time_report_count_csv:
            time_report_count_csv[d.strftime("%d-%m-%Y")] = {"0-1": 0, "1-2": 0, "2-3": 0,
                                                             "3-5": 0, "5-10": 0, "10-15": 0,
                                                             "15-60": 0}
        for i, val in enumerate(rows):
            if val not in time_report_count_csv[d.strftime("%d-%m-%Y")]:
                time_report_count_csv[d.strftime("%d-%m-%Y")][val] = 0
        dataAdd = [d.strftime("%d-%m-%Y"), str(time_report_count_csv[d.strftime("%d-%m-%Y")]["0-1"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["1-2"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["2-3"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["3-5"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["5-10"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["10-15"]),
                   str(time_report_count_csv[d.strftime("%d-%m-%Y")]["15-60"])]

        data.append(dataAdd)

    time_report_time = get_station_data(session['search_station'] + "_data.p")
    data.append([])
    data.append([])
    # data.append(["Operator Activity (From - To)"])
    lo_arr = []
    lo_arr_in = []
    # vo = 13
    # for x in range(7):
    #     d = now - timedelta(days=x)
    #     if d.strftime("%d-%m-%Y") not in time_report_time:
    #         time_report_time[d.strftime("%d-%m-%Y")] = {"1-2": [], "2-3": [],
    #                                                      "3-5": [], "5-10": [], "10-15": [],
    #                                                      "15-60": []}
    #
    #
    #     for i, val in enumerate(rows):
    #         if val not in time_report_time[d.strftime("%d-%m-%Y")]:
    #             time_report_time[d.strftime("%d-%m-%Y")][val] = []
    #     lo_arr.append(vo)
    #
    #
    #     data.append([d.strftime("%d-%m-%Y")])
    #     vo = vo + 1
    #     for o,datapop  in enumerate(rows):
    #         data.append([datapop +' Min'])
    #         lo_arr_in.append(vo)
    #         vo = vo + 1
    #         for i, val in enumerate(time_report_time[d.strftime("%d-%m-%Y")][datapop]):
    #             data.append([str(val["from"]) + " - " + str(val["to"]), str(val["diff"])])
    #             vo = vo + 1
    #         data.append([])
    #         vo = vo + 1

    fname2 = r'Shipping_Hrs' + '_' + session['_user_id'] + '.xlsx'
    workbook = openpyxl.Workbook(fname2)
    if os.path.isfile(fname2):
        print('old file')
    else:
        print('new file')
        workbook2 = openpyxl.Workbook(fname2)
        workbook2.save(fname2)
    wb = openpyxl.load_workbook(fname2)
    ws = wb.active

    sheet = wb.active

    for row in data:
        sheet.append(row)
    redFill = PatternFill(start_color='FFFF00',
                          end_color='FFFF99',
                          fill_type='solid')

    redFill_in = PatternFill(start_color='A9A9A9',
                             end_color='E8E8E8',
                             fill_type='solid')

    for cell in ws["2:2"]:
        cell.fill = redFill
        cell.font = Font(size="14", color='484848')
    for p, pon in enumerate(lo_arr):
        for cell in ws[str(pon) + ":" + str(pon)]:
            cell.fill = redFill
            cell.font = Font(size="14", color='484848')
    for p, pon in enumerate(lo_arr_in):
        for cell in ws[str(pon) + ":" + str(pon)]:
            cell.fill = redFill_in
            cell.font = Font(size="14", color='FFFFFF')

    fontStyle = Font(size="20", color='484848')
    ws['A1'].font = fontStyle

    ws['A12'].font = fontStyle

    wb.save(fname2)
    excelDownload = open(fname2, 'rb').read()
    os.remove(fname2)
    return Response(
        excelDownload,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition":
                     "attachment; filename=" + fname2})


@blueprint.route('/report-csv-op-activity')
@login_required
def report_csv_op_activity():
    now = datetime.now()

    rows = ["Date", "0 Hr", "1 Hr", "2 Hr", "3 Hr", "4 Hr", "5 Hr", "6 Hr", "7 Hr", "8 Hr", "9 Hr", "10 Hr", "11 Hr"
        , "12 Hr", "13 Hr", "14 Hr", "15 Hr", "16 Hr", "17 Hr", "18 Hr", "19 Hr", "20 Hr", "21 Hr", "22 Hr",
            "23 Hr"]
    time_report_count_csv = get_station_data(session['search_station'] + "_hrShippingCount.p")

    data = []
    data.append(["Shipping hourly report"])
    data.append(rows)
    for x in range(7):
        d = now - timedelta(days=x)
        print(d.strftime("%d-%m-%Y"))
        if d.strftime("%d-%m-%Y") not in time_report_count_csv:
            time_report_count_csv[d.strftime("%d-%m-%Y")] = {0: 0, 1: 0, 2: 0,
                                                             3: 0, 4: 0, 5: 0,
                                                             6: 0, 7: 0, 8: 0,
                                                             9: 0, 10: 0, 11: 0,
                                                             12: 0, 13: 0, 14: 0,
                                                             15: 0, 16: 0, 17: 0,
                                                             18: 0, 19: 0, 20: 0,
                                                             21: 0, 22: 0, 23: 0}
        for i in range(24):
            if i not in time_report_count_csv[d.strftime("%d-%m-%Y")]:
                time_report_count_csv[d.strftime("%d-%m-%Y")][i] = 0

        data.append([d.strftime("%d-%m-%Y"),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][0]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][1]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][2]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][3]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][4]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][5]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][6]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][7]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][8]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][9]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][10]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][11]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][12]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][13]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][14]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][15]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][16]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][17]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][18]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][19]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][20]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][21]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][22]),
                     str(time_report_count_csv[d.strftime("%d-%m-%Y")][23])])

    fname2 = r'Operator_Activity' + '_' + session['_user_id'] + '.xlsx'
    workbook = openpyxl.Workbook(fname2)
    if os.path.isfile(fname2):
        print('old file')
    else:
        print('new file')
        workbook2 = openpyxl.Workbook(fname2)
        workbook2.save(fname2)
    wb = openpyxl.load_workbook(fname2)
    ws = wb.active

    sheet = wb.active
    for row in data:
        sheet.append(row)

    redFill = PatternFill(start_color='FFFF00',
                          end_color='FFFF99',
                          fill_type='solid')

    redFill_in = PatternFill(start_color='A9A9A9',
                             end_color='E8E8E8',
                             fill_type='solid')

    for cell in ws["2:2"]:
        cell.fill = redFill
        cell.font = Font(size="14", color='484848')
    fontStyle = Font(size="20", color='484848')
    ws['A1'].font = fontStyle

    ws['A12'].font = fontStyle

    wb.save(fname2)
    excelDownload = open(fname2, 'rb').read()
    os.remove(fname2)
    return Response(
        excelDownload,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition":
                     "attachment; filename=" + fname2})


def is_between(time, time_range):
    if time_range[1] < time_range[0]:
        return time >= time_range[0] or time <= time_range[1]
    return time_range[0] <= time <= time_range[1]

@blueprint.route('/report/tags')
@login_required
def report_shipping_tags():
    stationTags = motions.actionin_shipping_data_tag(stationDis[session['search_station']])
     
    tags = {}
    for tag in stationTags:
        tags[tag[1]] = tag[0] 
    result = {"result":tags}
    return jsonify(result), 200
    


@blueprint.route('/report/shipping/customer')
@login_required
def report_shipping_customer():
    stationTime = motions.actionin_shipping_data(stationDis[session['search_station']])
    customers = []
    shipping_cus_count = {}
    warehouses = {}

    now = datetime.now()
    for x in range(8):
        d = now - timedelta(days=x)
        dt_objr = d.strftime("%d-%m-%Y")
        if dt_objr not in shipping_cus_count:
            shipping_cus_count[dt_objr] = {}

    for station in stationTime:
        stationstarttime = session['timedep'] + station[3]
        dt_obj = datetime.fromtimestamp(stationstarttime).strftime('%d-%m-%Y')
        str_customer = str(station[6])[0:3]
        wh = station[7].strip()
        if str_customer not in customers:
            customers.append(str_customer)
        if str_customer not in shipping_cus_count[dt_obj]:
            shipping_cus_count[dt_obj][str_customer] = 0
        shipping_cus_count[dt_obj][str_customer] = shipping_cus_count[dt_obj][str_customer] + 1
    data = []
    data.append(["Station Shipping Count - Customer"])
    daterange = []
    daterange.append('Customer')
    for x in range(7):
        d = now - timedelta(days=x)
        dt_objr = d.strftime("%d-%m-%Y")
        daterange.append(dt_objr)

    data.append(daterange)

    for customer in customers:
        sp = []
        sp.append(customer)
        for x in range(7):

            d = now - timedelta(days=x)
            dt_objr = d.strftime("%d-%m-%Y")
            if customer in shipping_cus_count[dt_objr]:
                sp.append(shipping_cus_count[dt_objr][customer])
            else:
                sp.append(0)

        data.append(sp)

    fname2 = r'Customer_Based_Shipping_Count' + '_' + session['_user_id'] + '.xlsx'
    workbook = openpyxl.Workbook(fname2)
    if os.path.isfile(fname2):
        print('old file')
    else:
        print('new file')
        workbook2 = openpyxl.Workbook(fname2)
        workbook2.save(fname2)
    wb = openpyxl.load_workbook(fname2)
    ws = wb.active

    sheet = wb.active
    for row in data:
        sheet.append(row)

    redFill = PatternFill(start_color='FFFF00',
                          end_color='FFFF99',
                          fill_type='solid')

    redFill_in = PatternFill(start_color='A9A9A9',
                             end_color='E8E8E8',
                             fill_type='solid')

    for cell in ws["2:2"]:
        cell.fill = redFill
        cell.font = Font(size="14", color='484848')
    fontStyle = Font(size="20", color='484848')
    ws['A1'].font = fontStyle

    ws['A12'].font = fontStyle

    wb.save(fname2)
    excelDownload = open(fname2, 'rb').read()
    os.remove(fname2)
    return Response(
        excelDownload,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition":
                     "attachment; filename=" + fname2})


@blueprint.route('/report')
@login_required
def report():
    stationTime = motions.motion_loader_byarea(session['search_station'])
    time_report_count = {}
    time_report = {}
    time_report_hrs = {}
    time_report_time = {}
    pretime = {}

    station_all_day = {}
    for station in stationTime:
        stationstarttime = session['timedep'] + station[0]
        stationendtime = session['timedep'] + station[4]
        current_time = datetime.fromtimestamp(stationstarttime).strftime('%H:%M:%S')

        dt_obj = datetime.fromtimestamp(stationstarttime).strftime('%d-%m-%Y')
        # print(  current_time)
        if is_between(current_time, ("07:00:00", "20:00:00")):
            if dt_obj not in time_report_count:
                time_report_count[dt_obj] = {"1-2": 0, "2-3": 0, "3-5": 0, "5-10": 0, "10-15": 0, "15-60": 0}
                time_report[dt_obj] = {"1-2": 0, "2-3": 0, "3-5": 0, "5-10": 0, "10-15": 0, "15-60": 0}
                time_report_hrs[dt_obj] = {"1-2": 0, "2-3": 0, "3-5": 0, "5-10": 0, "10-15": 0, "15-60": 0}
                time_report_time[dt_obj] = {"1-2": [], "2-3": [], "3-5": [], "5-10": [], "10-15": [], "15-60": []}
                pretime[dt_obj] = 0
            if pretime[dt_obj] > int(stationstarttime) or pretime[dt_obj] == 0:
                pretime[dt_obj] = int(stationstarttime)
            if 59 < station[5] <= 120:
                p_time = time_report[dt_obj]["1-2"]
                time_report_count[dt_obj]["1-2"] = int(time_report_count[dt_obj]["1-2"]) + 1
                time_report[dt_obj]["1-2"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["1-2"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["1-2"].append({"from": datetime.fromtimestamp(int(stationstarttime)),
                                                        "to": datetime.fromtimestamp(int(stationendtime)),
                                                        "diff": int(station[5])})
            if 120 < station[5] <= 180:
                p_time = time_report[dt_obj]["2-3"]
                time_report_count[dt_obj]["2-3"] = int(time_report_count[dt_obj]["2-3"]) + 1
                time_report[dt_obj]["2-3"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["2-3"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["2-3"].append({"from": datetime.fromtimestamp(int(stationstarttime)),
                                                        "to": datetime.fromtimestamp(int(stationendtime)),
                                                        "diff": int(station[5])})
            if 180 < station[5] <= 300:
                p_time = time_report[dt_obj]["3-5"]
                time_report_count[dt_obj]["3-5"] = int(time_report_count[dt_obj]["3-5"]) + 1
                time_report[dt_obj]["3-5"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["3-5"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["3-5"].append({"from": datetime.fromtimestamp(int(stationstarttime)),
                                                        "to": datetime.fromtimestamp(int(stationendtime)),
                                                        "diff": int(station[5])})

            if 300 < station[5] <= 600:
                p_time = time_report[dt_obj]["5-10"]
                time_report_count[dt_obj]["5-10"] = int(time_report_count[dt_obj]["5-10"]) + 1
                time_report[dt_obj]["5-10"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["5-10"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["5-10"].append({"from": datetime.fromtimestamp(int(stationstarttime)),
                                                         "to": datetime.fromtimestamp(int(stationendtime)),
                                                         "diff": int(station[5])})

            if 600 < station[5] <= 900:
                p_time = time_report[dt_obj]["10-15"]
                time_report_count[dt_obj]["10-15"] = int(time_report_count[dt_obj]["10-15"]) + 1
                time_report[dt_obj]["10-15"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["10-15"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["10-15"].append({"from": datetime.fromtimestamp(int(stationstarttime)),
                                                          "to": datetime.fromtimestamp(int(stationendtime)),
                                                          "diff": int(station[5])})

            if 900 < station[5] <= 1000:
                p_time = time_report[dt_obj]["15-60"]
                time_report_count[dt_obj]["15-60"] = int(time_report_count[dt_obj]["15-60"]) + 1
                time_report[dt_obj]["15-60"] = (int(p_time) + station[5])
                time_report_hrs[dt_obj]["15-60"] = convert_time((int(p_time) + station[5]))
                time_report_time[dt_obj]["15-60"].append({"from": datetime.fromtimestamp(int(stationstarttime)),
                                                          "to": datetime.fromtimestamp(int(stationendtime)),
                                                          "diff": int(station[5])})

    # print(time_report_time)
    station_all_day = {"time_report_count": time_report_count, "time_report": time_report,
                       "time_report_hrs": time_report_hrs, "pretime": pretime}

    result = {'result': station_all_day}

    pickle.dump(time_report_time, open(get_correct_path(session['search_station'] + "_data.p"), "wb"))
    pickle.dump(time_report_count, open(get_correct_path(session['search_station'] + "_data_count.p"), "wb"))
    return jsonify(result), 200





def get_station_data(name):
    try:
        data = pickle.load(open(get_correct_path(name), "rb"))
    except EOFError:
        data = list()
    return data


def get_correct_path(relative_path):
    p = os.path.abspath(".").replace('/dist', "")
    return os.path.join(p, relative_path)


def convert_time(seconds):
    return seconds / 60
    seconds = seconds % (24 * 3600)
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60

    return "%d:%02d:%02d" % (hour, minutes, seconds)
